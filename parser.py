"""
parser.py — Extract monthly AI-tool adoption data from the Excel workbook.

Each tool sheet contains vertically stacked month blocks.
Two sub-tables per block:
  1. By-Leader  (one row per manager)
  2. By-Function (one row per function, mapped to a leader)

We extract the *Function-level* sub-table plus the Total row for each month.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_values(row) -> list[Any]:
    """Return plain Python values for an openpyxl row tuple."""
    return [cell.value for cell in row]


def _parse_month_label(text: str | None) -> str | None:
    """Normalise a month label like 'Feb. 26' or 'Jan. 26'."""
    if not text or not isinstance(text, str):
        return None
    text = text.strip()
    # Accept patterns like "Feb. 26", "Jan. 26", "Dec. 24 - Nov. 25"
    if re.match(r"^[A-Za-z]{3,}\.\s*\d{2}$", text):
        return text
    if re.match(r"^[A-Za-z]{3,}\.\s*\d{2}\s*-\s*[A-Za-z]{3,}\.\s*\d{2}$", text):
        return text
    return None


def _month_sort_key(label: str) -> tuple[int, int]:
    """Return (year, month_index) for sorting chronologically.

    Handles 'Feb. 26'  -> (2026, 2)
              'Jan. 26' -> (2026, 1)
              'Dec. 24 - Nov. 25' -> uses the *end* date -> (2025, 11)
    """
    month_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    }
    # If range like "Dec. 24 - Nov. 25", use the end date
    range_match = re.match(
        r"[A-Za-z]{3,}\.\s*\d{2}\s*-\s*([A-Za-z]{3,})\.\s*(\d{2})", label
    )
    if range_match:
        mon_str = range_match.group(1).lower()[:3]
        yr = int(range_match.group(2)) + 2000
        return (yr, month_map.get(mon_str, 0))

    single_match = re.match(r"([A-Za-z]{3,})\.\s*(\d{2})", label)
    if single_match:
        mon_str = single_match.group(1).lower()[:3]
        yr = int(single_match.group(2)) + 2000
        return (yr, month_map.get(mon_str, 0))

    return (0, 0)


# ---------------------------------------------------------------------------
# Sheet-specific parsers
# ---------------------------------------------------------------------------

def _parse_standard_sheet(ws, tool_name: str) -> list[dict]:
    """Parse AI Code or NABU sheets.

    Layout per month block:
        Row x:   [month_label, None, ...]
        Row x+1: header row (Leader-level): None, None, 'Leader', 'Total', ...
        ...rows...
        Row y:   header row (Function-level): None, 'Function', 'Leader', 'Total', ...
        ...rows...

    We want the Function-level rows (column B = Function) and the Total row.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False))
    records: list[dict] = []

    current_month: str | None = None
    in_function_block = False

    for row in rows:
        vals = _cell_values(row)

        # Detect month label (column A, other cells mostly None)
        month_candidate = _parse_month_label(vals[0])
        if month_candidate:
            current_month = month_candidate
            in_function_block = False
            continue

        if current_month is None:
            continue

        # Detect headers — function-level header has 'Function' in col B (index 1)
        if vals[1] == "Function":
            in_function_block = True
            continue

        # Detect leader-level header — has 'Leader' in col C but no 'Function' in col B
        if vals[2] == "Leader" and vals[1] is None:
            in_function_block = False
            continue

        if not in_function_block:
            continue

        # Data row inside function block
        function = vals[1]
        leader = vals[2]
        total_hc = vals[3]
        adoption_count = vals[4]
        active_count = vals[5]
        adoption_rate = vals[6]
        active_rate = vals[7]

        if leader is None and function is None:
            continue

        label = str(leader) if leader else str(function)
        is_total = label.lower().strip() == "total"

        records.append({
            "Month": current_month,
            "Tool": tool_name,
            "Function": "Total" if is_total else str(function) if function else "",
            "Leader": str(leader) if leader else "",
            "Headcount": total_hc,
            "Adoption Count": adoption_count,
            "Active Count": active_count,
            "Adoption Rate": adoption_rate,
            "Active Rate": active_rate,
        })

    return records


def _parse_copilot_sheet(ws) -> list[dict]:
    """Parse Microsoft Copilot sheet — function-level summary blocks only.

    The sheet has multiple month blocks, each with a function-level summary.
    Feb. 26 function block has columns:
        Function, Leader, Total, Active Copilot users, Avg weekly actions,
        Copilot licensed users, Total actions, Adoption Rate, Active Rate,
        Avg. Action Per User
    Older blocks may have different column order.
    We detect the header row and map columns dynamically.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False))
    records: list[dict] = []

    current_month: str | None = None

    # Column name normalisations
    col_aliases = {
        "function": "Function",
        "leader": "Leader",
        "total": "Total",
        "active copilot users": "Active Count",
        "average weekly actions performed": "Avg Weekly Actions",
        "copilot licensed users": "Licensed Users",
        "total actions performed": "Total Actions",
        "adoption rate": "Adoption Rate",
        "active rate": "Active Rate",
        "avg. action per user": "Avg Action Per User",
        "nabu adoption": "Adoption Count",
        "nabu active": "Active Count",
    }

    header_map: dict[int, str] | None = None
    in_function_block = False

    for row in rows:
        vals = _cell_values(row)

        # Detect month label
        month_candidate = _parse_month_label(vals[0])
        if month_candidate:
            current_month = month_candidate
            in_function_block = False
            header_map = None
            continue

        if current_month is None:
            continue

        # Detect a function-level header: col B = 'Function'
        if vals[1] == "Function":
            in_function_block = True
            header_map = {}
            for i, v in enumerate(vals):
                if v and isinstance(v, str):
                    key = v.strip().lower()
                    mapped = col_aliases.get(key, v.strip())
                    header_map[i] = mapped
            continue

        # Detect organization-level header (skip this block)
        if vals[1] == "Organization":
            in_function_block = False
            header_map = None
            continue

        if not in_function_block or header_map is None:
            continue

        # Build a record from the current data row
        row_dict: dict[str, Any] = {}
        for i, mapped_name in header_map.items():
            if i < len(vals):
                row_dict[mapped_name] = vals[i]

        function = row_dict.get("Function")
        leader = row_dict.get("Leader")
        if function is None and leader is None:
            continue

        is_total = (leader and str(leader).strip().lower() == "total") or (
            function is None and leader and str(leader).strip().lower() == "total"
        )

        records.append({
            "Month": current_month,
            "Tool": "Microsoft Copilot",
            "Function": "Total" if is_total else str(function) if function else "",
            "Leader": str(leader) if leader else "",
            "Headcount": row_dict.get("Total"),
            "Adoption Count": row_dict.get("Licensed Users"),  # licensed ≈ adopted
            "Active Count": row_dict.get("Active Count"),
            "Adoption Rate": row_dict.get("Adoption Rate"),
            "Active Rate": row_dict.get("Active Rate"),
            "Avg Weekly Actions": row_dict.get("Avg Weekly Actions"),
            "Total Actions": row_dict.get("Total Actions"),
            "Avg Action Per User": row_dict.get("Avg Action Per User"),
        })

    return records


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

# Sheets we know how to parse and the parser to use.
# Add new tool sheets here as they appear in the workbook.
TOOL_SHEETS: dict[str, str] = {
    "AI Code": "standard",
    "NABU": "standard",
    "Microsoft Copilot": "copilot",
}


def parse_workbook(path: str | Path) -> pd.DataFrame:
    """Parse the workbook and return a tidy DataFrame with all months / tools / functions."""
    wb = openpyxl.load_workbook(str(path), data_only=True)

    all_records: list[dict] = []

    for sheet_name, parser_type in TOOL_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [skip] Sheet '{sheet_name}' not found in workbook.")
            continue
        ws = wb[sheet_name]
        if parser_type == "standard":
            records = _parse_standard_sheet(ws, sheet_name)
        elif parser_type == "copilot":
            records = _parse_copilot_sheet(ws)
        else:
            continue
        print(f"  Parsed '{sheet_name}': {len(records)} function-level records")
        all_records.extend(records)

    df = pd.DataFrame(all_records)

    if df.empty:
        return df

    # Convert rate columns to float
    for col in ["Adoption Rate", "Active Rate"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Add sort key for chronological ordering
    df["_sort"] = df["Month"].apply(_month_sort_key)
    df = df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

    return df
